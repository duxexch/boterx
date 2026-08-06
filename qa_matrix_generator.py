#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
QA Matrix Generator — generates professional iGaming-level test cases
for any game using the Matrix Approach (Core × Environmental matrices).

Usage:
    python qa_matrix_generator.py

Output:
    qa_exports/Aviator_QA_TestCases.xlsx (400+ cases)
    qa_exports/Crash_QA_TestCases.xlsx (300+ cases)
    ... one Excel per game

Pattern: 40-50 Core Functional cases × Network/Telegram/Security/UI/Perf matrices
= 300-500 unique test cases per game, zero duplication.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# ENVIRONMENTAL MATRICES (reusable across all games)
# ============================================================

NETWORK_MATRIX = [
    {"suffix": "during Offline", "type": "Network", "steps": "1. Disconnect network\n2. Perform action", "expected": "Action is queued or rejected gracefully. 'Network Disconnected' overlay shown. No data corruption", "priority": "High", "severity": "High"},
    {"suffix": "during Slow 3G", "type": "Network", "steps": "1. Throttle to Slow 3G (400ms latency)\n2. Perform action", "expected": "Action completes with delay. Animation doesn't freeze. No timeout errors", "priority": "Medium", "severity": "Medium"},
    {"suffix": "during Packet Loss", "type": "Network", "steps": "1. Set 30% packet loss\n2. Perform action", "expected": "Auto-retry or graceful degradation. SSE reconnects. No stuck state", "priority": "High", "severity": "High"},
    {"suffix": "after Reconnect", "type": "Network", "steps": "1. Disconnect\n2. Wait 5s\n3. Reconnect", "expected": "Game resyncs to current state. Bet/round status preserved. Multiplier jumps to server value", "priority": "Critical", "severity": "Critical"},
    {"suffix": "during Timeout", "type": "Network", "steps": "1. Set server timeout to 1s\n2. Perform action", "expected": "Client shows 'Request timeout' toast. Retry button appears. No silent failure", "priority": "High", "severity": "High"},
]

TELEGRAM_MATRIX = [
    {"suffix": "via TMA MainButton", "type": "Telegram Mini App", "steps": "1. Use Telegram MainButton for action", "expected": "Action triggers correctly via TMA SDK. Button text updates to next state", "priority": "High", "severity": "High"},
    {"suffix": "via TMA BackButton", "type": "Telegram Mini App", "steps": "1. Click TMA BackButton during action", "expected": "Confirm exit dialog shown, or action cancelled. NOT browser default back", "priority": "High", "severity": "High"},
    {"suffix": "after TMA Theme Change", "type": "Telegram Mini App", "steps": "1. Change TMA theme mid-action\n2. Observe UI", "expected": "Colors update without breaking state. Game continues without reset", "priority": "Medium", "severity": "Medium"},
    {"suffix": "after TMA Resume from Background", "type": "Telegram Mini App", "steps": "1. Minimize TMA\n2. Wait 30s\n3. Resume", "expected": "Game resyncs to current round state. No stale multiplier. Bet status preserved", "priority": "Critical", "severity": "Critical"},
    {"suffix": "within TMA Safe Area", "type": "Telegram Mini App", "steps": "1. Open on notched device\n2. Check all UI elements", "expected": "No element overlaps notch or TMA header/footer. Bottom row visible above MainButton", "priority": "Medium", "severity": "Medium"},
]

SECURITY_MATRIX = [
    {"suffix": "with Replay Attack", "type": "Security", "steps": "1. Intercept action packet\n2. Replay it 3 times", "expected": "Server rejects duplicates (409 Conflict). Only first action processed. Balance unchanged after 1st", "priority": "Critical", "severity": "Critical"},
    {"suffix": "with Invalid JWT", "type": "Security", "steps": "1. Expire/corrupt session token\n2. Perform action", "expected": "401 Unauthorized. Game redirects to re-auth. No balance change", "priority": "Critical", "severity": "Critical"},
    {"suffix": "after Session Expired", "type": "Security", "steps": "1. Wait for session to expire\n2. Perform action", "expected": "Session refresh or re-login prompt. No silent acceptance of expired session", "priority": "High", "severity": "High"},
    {"suffix": "with Rate Limit Exceeded", "type": "Security", "steps": "1. Send 20 requests in 1s\n2. Perform action on 21st", "expected": "429 Too Many Requests. 'طلبات كثيرة' toast shown. No DoS", "priority": "High", "severity": "High"},
    {"suffix": "with Duplicate Payload", "type": "Security", "steps": "1. Send same request_id twice", "expected": "Second request rejected (409). request_id dedup works correctly", "priority": "Critical", "severity": "Critical"},
]

UI_MATRIX = [
    {"suffix": "in Landscape", "type": "UI", "steps": "1. Rotate device to landscape\n2. Perform action", "expected": "Layout scales. No truncation. Canvas/game area fills screen", "priority": "Medium", "severity": "Medium"},
    {"suffix": "in Dark Mode", "type": "UI", "steps": "1. Switch to dark theme\n2. Perform action", "expected": "Colors update. Text readable. No white-on-white or invisible elements", "priority": "Medium", "severity": "Medium"},
    {"suffix": "in RTL (Arabic)", "type": "UI", "steps": "1. Set language to Arabic\n2. Perform action", "expected": "Layout is RTL. Text direction correct. No LTR artifacts. Numbers LTR acceptable", "priority": "Medium", "severity": "Medium"},
    {"suffix": "on Low-End Device", "type": "UI", "steps": "1. Throttle CPU 4x slowdown\n2. Perform action", "expected": "FPS >= 30. No freeze. Animation degrades gracefully (reduces particles)", "priority": "High", "severity": "High"},
    {"suffix": "on Small Screen (320px)", "type": "UI", "steps": "1. Set viewport to 320px width\n2. Perform action", "expected": "All buttons visible. No horizontal scroll. Bet input accessible. Canvas scales", "priority": "Medium", "severity": "Medium"},
]

PERFORMANCE_MATRIX = [
    {"suffix": "with Low Memory", "type": "Performance", "steps": "1. Simulate low memory\n2. Perform 50 rounds", "expected": "No OOM crash. GC kicks in. Game recovers. Memory stable after 50 rounds", "priority": "High", "severity": "High"},
    {"suffix": "with 100 concurrent players (SSE)", "type": "Performance", "steps": "1. Connect 100 SSE clients\n2. Play 10 rounds", "expected": "Server handles within 2GB/1core. SSE delivers <2s latency. No 502/503", "priority": "Critical", "severity": "Critical"},
]

ALL_MATRICES = NETWORK_MATRIX + TELEGRAM_MATRIX + SECURITY_MATRIX + UI_MATRIX + PERFORMANCE_MATRIX

# ============================================================
# GAME-SPECIFIC CORE TEST CASES
# ============================================================

AVIATOR_CORE = [
    # --- Round Lifecycle ---
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify round starts with 6s countdown", "pre": "User in game, no active round", "steps": "1. Wait for previous round to end\n2. Observe countdown", "expected": "Countdown starts at 6, decrements each second. 'ضع رهانك' message shown", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify multiplier starts at 1.00x and increments", "pre": "Countdown reached 0", "steps": "1. Observe multiplier after countdown", "expected": "Multiplier starts at 1.00x, increments smoothly. Plane takes off at 22° angle", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify slow growth phase (first 2s on runway)", "pre": "Round started", "steps": "1. Observe multiplier for first 2 seconds", "expected": "Multiplier increments slowly (GROWTH=1.003). Numbers move slowly. Plane on runway", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify fast growth phase (after 2s)", "pre": "2 seconds elapsed", "steps": "1. Observe multiplier after 2s mark", "expected": "Growth rate increases (GROWTH=1.012). Multiplier accelerates. Plane climbs", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Crash", "title": "Verify crash ends the round", "pre": "Round in progress", "steps": "1. Wait for crash point\n2. Observe crash", "expected": "💥 message shown. Crash point displayed. Plane disappears. All un-cashed bets lost", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Crash", "title": "Verify crash shows total distributed profits", "pre": "Round crashed", "steps": "1. Read post-crash message", "expected": "'💰 إجمالي الأرباح الموزعة: X — Y لاعب ربح من Z' shown for 5 seconds", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify new round starts automatically after crash", "pre": "Crash results shown", "steps": "1. Wait 5 seconds after crash", "expected": "New countdown begins at 6. Plane returns to runway. History bar updated", "priority": "Critical", "severity": "Critical"},

    # --- Betting ---
    {"module": "Betting", "feature": "Place Bet", "title": "Verify placing valid bet during countdown", "pre": "Balance=100, countdown active", "steps": "1. Enter 10 in bet field\n2. Click Bet button", "expected": "10 deducted. Button shows 'إلغاء'. Bet appears in my bets panel. Balance=90", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Cancel Bet", "title": "Verify cancelling bet before takeoff", "pre": "Bet placed, countdown active", "steps": "1. Click 'إلغاء' button", "expected": "Bet cancelled. 10 refunded. Button reverts to 'رهان'. Balance=100", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify bet rejected after countdown ends", "pre": "Countdown=0, plane flying", "steps": "1. Click Bet button", "expected": "Toast: '⏳ انتظر نهاية الجولة الحالية!'. No bet placed. Balance unchanged", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify insufficient balance shows deposit modal", "pre": "Balance=0", "steps": "1. Enter 10 in bet field\n2. Click Bet", "expected": "VEX deposit modal appears. No bet placed. No crash", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Dual Slots", "title": "Verify two simultaneous bets (slot 1 + slot 2)", "pre": "Balance=100", "steps": "1. Enable slot 2\n2. Bet 10 on slot 1\n3. Bet 20 on slot 2", "expected": "Both bets accepted. 30 deducted. Both appear in my bets. Balance=70", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Dual Slots", "title": "Verify slot 2 toggle disables/enables", "pre": "Slot 2 disabled", "steps": "1. Toggle slot 2 on\n2. Toggle slot 2 off", "expected": "Button enables/disables. Input field shows/hides. No orphan state", "priority": "Medium", "severity": "Medium"},

    # --- Cash Out ---
    {"module": "Betting", "feature": "Cash Out", "title": "Verify manual cash out at current multiplier", "pre": "Active bet=10, multiplier=2.50x", "steps": "1. Click 'سحب' button", "expected": "Win=25 (10×2.50). Balance updates. Button shows '✓ 25'. 🪂 parachute with name+amount", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Cash Out", "title": "Verify bet button shows increasing amount during flight", "pre": "Active bet=10, multiplier rising", "steps": "1. Observe button text during flight", "expected": "Button shows '💰 12' → '💰 18' → '💰 25' — updates with multiplier", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Auto Cash Out", "title": "Verify auto cash out at target multiplier", "pre": "Active bet=10, Auto Cashout=5x", "steps": "1. Start round\n2. Wait for 5.00x", "expected": "Cash out auto-triggers at exactly 5.00x. Payout=50. Parachute appears", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Cash Out", "title": "Verify cash out rejected if already crashed", "pre": "Multiplier crashed at 2x", "steps": "1. Click 'سحب' after crash", "expected": "Request rejected by server. 'انفجرت الطائرة' message. No payout. Balance unchanged", "priority": "Critical", "severity": "Critical"},

    # --- Provably Fair ---
    {"module": "Provably Fair", "feature": "Seed Hash", "title": "Verify seed_hash is broadcast before betting window", "pre": "New round starting", "steps": "1. Observe waiting message\n2. Check for seed_hash field", "expected": "seed_hash (64-char hex) present in SSE waiting message BEFORE bets open", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Server Seed Reveal", "title": "Verify server_seed revealed after crash", "pre": "Round crashed", "steps": "1. Read crash SSE message\n2. Check for server_seed field", "expected": "server_seed (64-char hex) present in crash message. SHA256(server_seed) matches seed_hash", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Verification", "title": "Verify crash point matches HMAC-SHA256 calculation", "pre": "server_seed + client_seed revealed", "steps": "1. Calculate HMAC-SHA256(server_seed, client_seed:round_id)\n2. Convert to crash point\n3. Compare with displayed crash", "expected": "Calculated crash point matches server-announced crash point exactly", "priority": "Critical", "severity": "Critical"},

    # --- UI / Animation ---
    {"module": "UI", "feature": "Plane Animation", "title": "Verify plane takes off at 22° angle", "pre": "Round started", "steps": "1. Observe plane trajectory for first 3.5s", "expected": "Plane moves from bottom-left to center at ~22° angle, then hovers. Background scrolls", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Weather", "title": "Verify random weather each round", "pre": "Multiple rounds played", "steps": "1. Play 5 rounds\n2. Observe sky background", "expected": "Sky changes randomly: day (blue+clouds), night (stars+moon), storm (rain+lightning)", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Parachute", "title": "Verify parachute shows player name + amount", "pre": "Player cashed out", "steps": "1. Observe parachute after cashout", "expected": "🪂 icon + gold pill with '+250' + player name below. Disappears after 3s", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection indicator updates", "pre": "Game loaded", "steps": "1. Observe topbar\n2. Disconnect network\n3. Reconnect", "expected": "🟢 connected → 🔴 disconnected → 🟡 reconnecting → 🟢 connected", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history bar updates after each round", "pre": "Round completed", "steps": "1. Observe history bar", "expected": "New crash point added to left. Color: green if ≥1.5x, red if <1.5x. Max 15 items", "priority": "Medium", "severity": "Medium"},

    # --- Wallet Sync ---
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance updates after bet", "pre": "Balance=100", "steps": "1. Place bet=10", "expected": "Balance pill flashes red, updates to 90 immediately", "priority": "High", "severity": "High"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance updates after cashout", "pre": "Active bet, multiplier=3x", "steps": "1. Cash out", "expected": "Balance pill flashes green, updates with payout. Matches server balance_after", "priority": "High", "severity": "High"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance polls every 5 seconds", "pre": "Idle (no bet)", "steps": "1. Admin approves deposit\n2. Wait 5s", "expected": "Balance updates to new value within 5s without manual refresh", "priority": "Medium", "severity": "Medium"},

    # --- Edge Cases ---
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify instant crash at 1.00x", "pre": "Round started", "steps": "1. Observe if crash_point=1.00", "expected": "Plane takes off and immediately crashes at 1.00x. All bets lost. History shows 1.00x", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify very high crash point (50x+)", "pre": "Round started", "steps": "1. Wait for high multiplier\n2. Observe", "expected": "Multiplier continues. No cap. Background darkens. Stars appear. No performance drop", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet amount=0 rejected", "pre": "Countdown active", "steps": "1. Enter 0 in bet field\n2. Click Bet", "expected": "No action taken. Button does not submit. No deduction", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet exceeds max (5000) rejected", "pre": "Countdown active", "steps": "1. Enter 99999 in bet field\n2. Click Bet", "expected": "Bet rejected or capped at 5000. Error message shown", "priority": "Medium", "severity": "Medium"},

    # --- Race Conditions ---
    {"module": "Betting", "feature": "Race Condition", "title": "Verify double-clicking Bet button rapidly", "pre": "Countdown active, Balance=100", "steps": "1. Click Bet 5 times in 100ms", "expected": "Only 1 bet placed (server dedup via request_id). Balance=90, not 50", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify double-clicking Cash Out rapidly", "pre": "Active bet, multiplier=3x", "steps": "1. Click Cash Out 5 times in 100ms", "expected": "Only 1 cashout processed. No double payout. Server rejects duplicates (409)", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify bet at exact timer=0 boundary", "pre": "Countdown at 0.01s", "steps": "1. Click Bet as timer hits 0", "expected": "Server uses server_ts to decide. If within window: accepted. If late: rejected with 'انتهت نافذة الرهان'", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify cashout at exact crash boundary", "pre": "Multiplier approaching crash_point", "steps": "1. Click Cash Out at same moment as crash", "expected": "Server checks mult vs crash_point. If mult < crash: win. If mult ≥ crash: rejected", "priority": "Critical", "severity": "Critical"},

    # --- Error Handling ---
    {"module": "Error Handling", "feature": "Server Error", "title": "Verify behavior when server returns 500", "pre": "Game in progress", "steps": "1. Simulate 500 error\n2. Observe UI", "expected": "Error toast shown. Game continues with last known state. No crash. Retry on next SSE", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "SSE Disconnect", "title": "Verify behavior when SSE stream drops", "pre": "Game in progress", "steps": "1. Kill SSE connection\n2. Observe", "expected": "Indicator turns 🔴. Auto-reconnect after 2s. State fetches from /api/aviator/state", "priority": "High", "severity": "High"},

    # --- Audit Trail ---
    {"module": "Database", "feature": "Audit Trail", "title": "Verify round logged in aviator_rounds table", "pre": "Round completed", "steps": "1. Query SQLite aviator_rounds table\n2. Check latest entry", "expected": "round_id, crash_point, seed_hash, client_seed, server_seed, bet_count, totals all present", "priority": "Medium", "severity": "Medium"},

    # --- Multi-Device ---
    {"module": "Multi-Device", "feature": "Sync", "title": "Verify same round visible on 2 devices", "pre": "2 devices logged in", "steps": "1. Open Aviator on Device A\n2. Open Aviator on Device B\n3. Observe multiplier", "expected": "Both show same multiplier at same time. Same crash point. Same countdown", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Bet Independence", "title": "Verify bet on Device A doesn't affect Device B", "pre": "Same user on 2 devices", "steps": "1. Place bet on Device A\n2. Check Device B", "expected": "Device B shows no active bet. Each device has independent bet state", "priority": "Medium", "severity": "Medium"},
]

# ============================================================
# CRASH — Core Test Cases (simplified Aviator, rocket graph)
# ============================================================

CRASH_CORE = [
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify rocket launch and multiplier increment", "pre": "User in game", "steps": "1. Wait for round start\n2. Observe multiplier", "expected": "Rocket launches. Multiplier starts 1.00x, increments. Graph curve draws on canvas", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Crash", "title": "Verify crash point display and explosion", "pre": "Round in progress", "steps": "1. Wait for crash\n2. Observe", "expected": "💥 explosion on canvas. Crash point shown. Screen shake. All un-cashed bets lost", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Round Lifecycle", "title": "Verify auto-restart after crash", "pre": "Crash occurred", "steps": "1. Wait 5 seconds", "expected": "New countdown begins. Graph clears. History bar updates with crash point", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify placing bet during countdown", "pre": "Balance=100, countdown active", "steps": "1. Enter 10\n2. Click Bet", "expected": "10 deducted. Button shows 'إلغاء'. Bet visible in my bets panel", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Cancel Bet", "title": "Verify cancelling bet before launch", "pre": "Bet placed, countdown active", "steps": "1. Click 'إلغاء'", "expected": "10 refunded. Button reverts to 'رهان'", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Cash Out", "title": "Verify manual cash out", "pre": "Active bet=10, mult=2.5x", "steps": "1. Click Cash Out", "expected": "Win=25. Balance updates. 🪂 parachute with name+amount", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Auto Cashout", "title": "Verify auto cashout at target", "pre": "Bet=10, Auto=3x", "steps": "1. Wait for 3.00x", "expected": "Auto cashout at 3.00x. Payout=30", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Dual Slots", "title": "Verify two simultaneous bets", "pre": "Balance=100", "steps": "1. Enable slot 2\n2. Bet 10 on each", "expected": "Both accepted. 20 deducted. Both appear in panel", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Bet Rejection", "title": "Verify bet rejected during flight", "pre": "Rocket flying", "steps": "1. Click Bet", "expected": "Toast: 'انتظر الجولة القادمة'. No bet placed", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal on low balance", "pre": "Balance=0", "steps": "1. Enter 10\n2. Click Bet", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Graph", "title": "Verify rocket graph curve renders correctly", "pre": "Round started", "steps": "1. Observe canvas graph", "expected": "Green curve with gradient fill. Rocket icon at curve tip. Grid lines visible", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history updates after each round", "pre": "Round completed", "steps": "1. Observe top history bar", "expected": "New crash point added. Color-coded (green ≥1.5x, red <1.5x). Max 15 items", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Quick Stats", "title": "Verify quick stats panel (max, avg, rounds)", "pre": "Multiple rounds played", "steps": "1. Observe stats panel", "expected": "Shows highest crash, average, total rounds. Updates in real-time", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status icon", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "Provably Fair", "feature": "Seed Hash", "title": "Verify seed_hash broadcast before betting", "pre": "New round", "steps": "1. Check SSE waiting message", "expected": "seed_hash present before bets open", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Server Seed Reveal", "title": "Verify server_seed after crash", "pre": "Round crashed", "steps": "1. Check crash message", "expected": "server_seed revealed. SHA256 matches seed_hash", "priority": "Critical", "severity": "Critical"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance updates after bet and cashout", "pre": "Balance=100", "steps": "1. Bet 10\n2. Cashout at 2x", "expected": "After bet: 90 (red flash). After cashout: 110 (green flash)", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify instant crash at 1.00x", "pre": "Round started", "steps": "1. Observe crash_point=1.00", "expected": "Rocket launches and immediately crashes. All bets lost", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify double-click Cash Out", "pre": "Active bet, mult=3x", "steps": "1. Spam click 5x in 100ms", "expected": "Only 1 cashout. No double payout. Server dedup via request_id", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify bet at timer=0 boundary", "pre": "Timer at 0.01s", "steps": "1. Click Bet at 0", "expected": "Server_ts decides. Accepted if within window, rejected if late", "priority": "Critical", "severity": "Critical"},
    {"module": "Error Handling", "feature": "SSE Drop", "title": "Verify auto-reconnect after SSE drop", "pre": "Round in progress", "steps": "1. Kill SSE\n2. Wait 2s", "expected": "Auto-reconnect. State fetched from /api/aviator/state. No data loss", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Sync", "title": "Verify same round on 2 devices", "pre": "2 devices", "steps": "1. Open on both\n2. Observe", "expected": "Same multiplier, same crash, same countdown", "priority": "High", "severity": "High"},
    {"module": "Database", "feature": "Audit Trail", "title": "Verify round logged in DB", "pre": "Round completed", "steps": "1. Query aviator_rounds", "expected": "round_id, crash_point, seeds, bet_count, totals all present", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Total Distributed", "title": "Verify total profits shown after crash", "pre": "Crash occurred", "steps": "1. Read post-crash message", "expected": "'إجمالي الأرباح الموزعة: X — Y لاعب ربح من Z' for 5s", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet=0 rejected", "pre": "Countdown active", "steps": "1. Enter 0\n2. Click Bet", "expected": "No action. No deduction", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet exceeds max rejected", "pre": "Countdown active", "steps": "1. Enter 99999\n2. Click Bet", "expected": "Rejected or capped. Error shown", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify high crash (50x+)", "pre": "Round started", "steps": "1. Wait for 50x+", "expected": "No cap. Graph continues. Background darkens. No performance drop", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Bet Button Dynamic", "title": "Verify button shows increasing amount during flight", "pre": "Active bet=10", "steps": "1. Observe button during flight", "expected": "Shows '💰 12' → '💰 18' → updates with multiplier", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Parachute", "title": "Verify parachute with name+amount", "pre": "Cashout occurred", "steps": "1. Observe parachute", "expected": "🪂 + gold pill with amount + player name. Disappears after 3s", "priority": "Medium", "severity": "Medium"},
    {"module": "Error Handling", "feature": "Server 500", "title": "Verify behavior on server error 500", "pre": "Game in progress", "steps": "1. Simulate 500\n2. Observe", "expected": "Error toast. Game continues with last state. Retry on next SSE", "priority": "High", "severity": "High"},
]

# ============================================================
# MINES — Core Test Cases (grid reveal, mine avoidance)
# ============================================================

MINES_CORE = [
    {"module": "Game Engine", "feature": "Grid Setup", "title": "Verify 5x5 grid initializes with correct mine count", "pre": "Mines=5, bet placed", "steps": "1. Set mines to 5\n2. Click Start", "expected": "5x5 grid appears. 5 mines hidden. Server hash generated. Bet deducted", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Reveal Tile", "title": "Verify revealing a safe tile shows diamond", "pre": "Game started, bet=10", "steps": "1. Click a closed tile", "expected": "Tile flips with 💎 animation. Multiplier increases. Cash out button enables", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Hit Mine", "title": "Verify hitting a mine ends game", "pre": "Game started", "steps": "1. Click a tile containing a mine", "expected": "Tile reveals 💣. Screen shake. Game over. Bet lost. All mines revealed", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Cash Out", "title": "Verify cashing out before hitting a mine", "pre": "3 safe tiles revealed", "steps": "1. Click Cash Out", "expected": "Game ends. Win = bet × current multiplier. Grid reveals all mines. Balance updates", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Mine Count", "title": "Verify changing mine count (3/5/7/10)", "pre": "Idle", "steps": "1. Set mines to 3\n2. Set to 5\n3. Set to 7\n4. Set to 10", "expected": "Selection highlights. Next game uses selected count. Higher count = higher multiplier per tile", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Multiplier", "title": "Verify multiplier increases with each safe reveal", "pre": "Game started, mines=5", "steps": "1. Reveal safe tile\n2. Check multiplier\n3. Reveal another\n4. Check again", "expected": "Multiplier increases after each safe reveal. Based on probability formula (safe/total remaining)", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Next Multiplier Preview", "title": "Verify next multiplier preview shown", "pre": "Game started", "steps": "1. Observe 'المضاعف التالي' field", "expected": "Shows the multiplier if next tile is safe. Updates after each reveal", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Auto Cashout", "title": "Verify auto cashout at target multiplier", "pre": "Auto cashout=2x, bet=10", "steps": "1. Start game\n2. Reveal tiles until mult reaches 2x", "expected": "Auto cashout triggers at 2x. Payout=20. Game ends", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify bet during waiting phase", "pre": "Balance=100", "steps": "1. Enter 10\n2. Click Bet", "expected": "10 deducted. Grid appears. Bet accepted", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal on low balance", "pre": "Balance=0", "steps": "1. Enter 10\n2. Click Bet", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Full Reveal", "title": "Verify all mines revealed after game ends", "pre": "Game over (mine hit)", "steps": "1. Observe grid after mine hit", "expected": "All 25 tiles reveal. Mines show 💣. Safe tiles show 💎. Sequentual reveal animation", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Auto-Bet", "title": "Verify auto-bet continues for N rounds", "pre": "Auto-bet=5, bet=10", "steps": "1. Enable auto-bet\n2. Play 5 rounds", "expected": "After each round ends, new round starts automatically. Counter decrements. Stops at 0", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Tile Animation", "title": "Verify tile flip animation", "pre": "Game started", "steps": "1. Click a safe tile", "expected": "Tile rotates (gemReveal animation 0.4s). Green glow. 💎 icon appears", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Mine Animation", "title": "Verify mine reveal animation", "pre": "Game started", "steps": "1. Click a mine tile", "expected": "Tile rotates (mineReveal 0.5s). Red glow. 💣 icon. Screen shake", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Countdown", "title": "Verify 3-2-1 countdown before game start", "pre": "Bet placed", "steps": "1. Observe countdown", "expected": "3-2-1 countdown displayed. Game starts after countdown. Sound tick each number", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history updates after each game", "pre": "Game completed", "steps": "1. Observe history bar", "expected": "Win/loss result added. Green for win, red for loss", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Quick Stats", "title": "Verify stats panel (best, rounds, wins)", "pre": "Multiple games played", "steps": "1. Observe stats", "expected": "Best multiplier, total rounds, total wins shown. Updates live", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify mines=24 (max safe tiles=1)", "pre": "Mines=24", "steps": "1. Start game\n2. Click any tile", "expected": "Only 1 safe tile. Very high multiplier. Wrong click = instant loss", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify mines=3 (easy mode)", "pre": "Mines=3", "steps": "1. Start game\n2. Reveal multiple tiles", "expected": "22 safe tiles. Low multiplier per tile. Easy to reveal many", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify clicking same tile twice rapidly", "pre": "Game started", "steps": "1. Click tile(1,1) twice in 10ms", "expected": "Only 1 reveal. No double multiplier. Server rejects second click", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify clicking 2 tiles simultaneously", "pre": "Game started", "steps": "1. Click tile(0,0) and tile(4,4) at same time", "expected": "Both tiles process. If both safe, multiplier applies. If one is mine, game ends correctly", "priority": "High", "severity": "High"},
    {"module": "Provably Fair", "feature": "Verification", "title": "Verify mine positions match server seed", "pre": "Game ended", "steps": "1. Get server seed\n2. Calculate mine positions\n3. Compare", "expected": "Calculated positions match revealed game. SHA256 verified", "priority": "Critical", "severity": "Critical"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance after win and loss", "pre": "Balance=100, bet=10", "steps": "1. Win at 2x\n2. Lose on mine", "expected": "After win: 120 (green flash). After loss: 90 (no flash, bet already deducted)", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "Server Error", "title": "Verify behavior on 500 during reveal", "pre": "Game in progress", "steps": "1. Click tile\n2. Simulate 500", "expected": "Error toast. Tile does not reveal. Game state preserved. Retry possible", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Independence", "title": "Verify game state is per-player", "pre": "2 devices", "steps": "1. Start game on Device A\n2. Check Device B", "expected": "Device B shows no game. Each device has independent grid state", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify bet=0 rejected", "pre": "Idle", "steps": "1. Enter 0\n2. Click Bet", "expected": "No action. No grid", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Bet Rejection", "title": "Verify bet rejected during active game", "pre": "Game in progress", "steps": "1. Click Bet again", "expected": "Rejected. 'انتظر نهاية الجولة' toast", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Responsive", "title": "Verify grid on small screen (320px)", "pre": "320px viewport", "steps": "1. Open game\n2. Observe grid", "expected": "5x5 grid fits. Tiles are touchable (min 44px). No horizontal scroll", "priority": "Medium", "severity": "Medium"},
]

# ============================================================
# PLINKO — Core Test Cases (ball physics, risk levels)
# ============================================================

PLINKO_CORE = [
    {"module": "Game Engine", "feature": "Drop Ball", "title": "Verify ball drops from center top", "pre": "Balance=100, bet=10", "steps": "1. Click Drop", "expected": "Ball spawns at top center. Falls. 10 deducted. Ball visible with glow trail", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Peg Collision", "title": "Verify ball bounces off pegs", "pre": "Ball dropping", "steps": "1. Observe ball hitting pegs", "expected": "Ball changes direction left/right on each peg. No clipping through pegs. Physics realistic", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Multiplier Slot", "title": "Verify ball lands in multiplier slot", "pre": "Ball reaching bottom", "steps": "1. Wait for ball to settle\n2. Check slot", "expected": "Ball lands in a slot. Win = bet × slot multiplier. Slot glows. Balance updates", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Risk Level", "title": "Verify risk level changes multipliers", "pre": "Idle", "steps": "1. Set Risk=Low\n2. Check slots\n3. Set Risk=High\n4. Check slots", "expected": "Low: safe multipliers (2.5x max). High: extreme edges (25x), <1x in center", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Row Count", "title": "Verify row count 8/12 affects board", "pre": "Idle", "steps": "1. Set rows=8\n2. Set rows=12", "expected": "Board changes. More rows = more pegs, more slots, longer ball path", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify bet and drop", "pre": "Balance=100", "steps": "1. Enter 10\n2. Click Drop", "expected": "10 deducted. Ball drops. Bet visible in panel", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Auto-Bet", "title": "Verify auto-bet for N rounds", "pre": "Auto-bet=5, bet=10", "steps": "1. Enable auto-bet\n2. Wait for 5 balls", "expected": "5 balls drop automatically. Counter decrements. Stops at 0. Each deducts 10", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal on low balance", "pre": "Balance=0", "steps": "1. Enter 10\n2. Click Drop", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Ball Trail", "title": "Verify ball has glowing trail", "pre": "Ball dropping", "steps": "1. Observe ball during fall", "expected": "Gold trail behind ball. Fades over time. Radial glow on ball", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Slot Highlight", "title": "Verify slot glows when ball lands", "pre": "Ball landing", "steps": "1. Wait for ball to settle", "expected": "Winning slot has glow animation (1.5s). Color: gold for win, red for <1x", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Board Layout", "title": "Verify pegs render correctly", "pre": "Game loaded", "steps": "1. Observe board", "expected": "Pegs arranged in triangle. 8 or 12 rows. Bottom row has multiplier labels", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history after each drop", "pre": "Drop completed", "steps": "1. Observe history", "expected": "Multiplier result added. Green for ≥1x, red for <1x", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Quick Stats", "title": "Verify stats panel", "pre": "Multiple drops", "steps": "1. Observe stats", "expected": "Best multiplier, total drops, total wins. Updates live", "priority": "Low", "severity": "Low"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify ball lands on slot border", "pre": "Ball at bottom", "steps": "1. Force ball to hit edge of peg", "expected": "Server seed determines final slot. UI snaps ball to correct server slot", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify <1x multiplier (loss slot)", "pre": "Risk=High", "steps": "1. Drop ball\n2. Ball lands in 0.2x slot", "expected": "Win = bet × 0.2. Player loses 80% of bet. Balance updates correctly", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify spamming Drop faster than balance updates", "pre": "Balance=10, bet=2", "steps": "1. Click Drop 10x in 500ms", "expected": "Only 5 drops succeed (balance limit). Rest ignored with 'Insufficient balance'", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Verification", "title": "Verify ball path determined by server", "pre": "Drop completed", "steps": "1. Get server seed\n2. Calculate path\n3. Compare", "expected": "Client animation ends at server-determined slot. No client-side prediction", "priority": "Critical", "severity": "Critical"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance after win and loss", "pre": "Balance=100, bet=10", "steps": "1. Win at 5x\n2. Lose at 0.2x", "expected": "After win: 140 (green). After loss: 92 (red flash)", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "Network", "title": "Verify drop right before offline", "pre": "Bet=10", "steps": "1. Click Drop\n2. Instantly go offline", "expected": "Ball animation plays (client). On reconnect, result fetched from server. Balance correct", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Independence", "title": "Verify per-player ball drops", "pre": "2 devices", "steps": "1. Drop on Device A\n2. Check Device B", "expected": "Device B shows no ball. Independent drops", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Bet Rejection", "title": "Verify bet=0 rejected", "pre": "Idle", "steps": "1. Enter 0\n2. Click Drop", "expected": "No action. No ball drop", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Bet Rejection", "title": "Verify bet exceeds max rejected", "pre": "Idle", "steps": "1. Enter 99999\n2. Click Drop", "expected": "Rejected or capped. Error shown", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Sound", "title": "Verify peg bounce sound", "pre": "Ball dropping", "steps": "1. Observe sound on each peg hit", "expected": "Beep on each peg collision. Pitch varies slightly. Sound toggle works", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Responsive", "title": "Verify board on 320px screen", "pre": "320px viewport", "steps": "1. Open game\n2. Observe board", "expected": "Board scales. Multiplier labels readable. Drop button accessible", "priority": "Medium", "severity": "Medium"},
]

# ============================================================
# LUCKY WHEEL — Core Test Cases (spinning wheel, segments)
# ============================================================

WHEEL_CORE = [
    {"module": "Game Engine", "feature": "Spin", "title": "Verify wheel spins and stops on a segment", "pre": "Balance=100, bet=10", "steps": "1. Click Spin", "expected": "Wheel rotates. Slows down. Stops on a segment. Result displayed. 10 deducted", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Multiplier", "title": "Verify multiplier applied correctly", "pre": "Spin result", "steps": "1. Check segment value\n2. Verify payout", "expected": "Win = bet × segment multiplier. If 0x: loss. If 2x: win=20. Balance updates", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Segments", "title": "Verify 8 segments with different multipliers", "pre": "Idle", "steps": "1. Observe wheel", "expected": "8 colored segments: 0x (red), 0.5x (blue), 1.5x (gold), 2x (green), 3x (purple)", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Animation", "title": "Verify spin animation is smooth", "pre": "Spin initiated", "steps": "1. Observe wheel rotation", "expected": "Smooth rotation. Decelerates naturally (ease-spring). 5s duration. No stuttering", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify bet and spin", "pre": "Balance=100", "steps": "1. Enter 10\n2. Click Spin", "expected": "10 deducted. Wheel spins. Result shown after animation", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Auto-Spin", "title": "Verify auto-spin for N rounds", "pre": "Auto-spin=5, bet=10", "steps": "1. Enable auto-spin\n2. Wait for 5 spins", "expected": "5 automatic spins. Counter decrements. Stops at 0", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal", "pre": "Balance=0", "steps": "1. Enter 10\n2. Click Spin", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Bet Rejection", "title": "Verify bet during active spin rejected", "pre": "Wheel spinning", "steps": "1. Click Spin again", "expected": "Rejected. 'انتظر' toast. No double spin", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Countdown", "title": "Verify 3-2-1 countdown before spin", "pre": "Bet placed", "steps": "1. Observe countdown", "expected": "3-2-1 displayed. Spin starts after countdown. Sound tick", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Result Display", "title": "Verify result shown after spin", "pre": "Spin completed", "steps": "1. Observe result", "expected": "Segment highlighted. Win/loss displayed. '🎉 +X' or '💔 0'", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history after each spin", "pre": "Spin completed", "steps": "1. Observe history", "expected": "Result added. Green for win, red for loss", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Quick Stats", "title": "Verify stats panel", "pre": "Multiple spins", "steps": "1. Observe stats", "expected": "Best, rounds, wins shown. Updates live", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify pointer lands on segment border", "pre": "Spin ending", "steps": "1. Observe pointer at segment boundary", "expected": "Server determines result. Pointer snaps to correct segment. No ambiguous result", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify 0x segment (total loss)", "pre": "Spin result = 0x", "steps": "1. Observe result", "expected": "'💔 0' shown. No payout. Bet lost. Balance already deducted", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify double-click Spin", "pre": "Idle", "steps": "1. Click Spin 3x in 100ms", "expected": "Only 1 spin. No double deduction. Button disables after 1st click", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Verification", "title": "Verify segment determined by server seed", "pre": "Spin completed", "steps": "1. Get server seed\n2. Calculate segment\n3. Compare", "expected": "Calculated segment matches displayed result", "priority": "Critical", "severity": "Critical"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance after win and loss", "pre": "Balance=100, bet=10", "steps": "1. Win at 2x\n2. Lose at 0x", "expected": "After win: 110 (green). After loss: 90 (no change, already deducted)", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "Server Error", "title": "Verify behavior on 500 during spin", "pre": "Wheel spinning", "steps": "1. Simulate 500\n2. Observe", "expected": "Animation continues (client). Result syncs on reconnect. No stuck state", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Independence", "title": "Verify per-player spins", "pre": "2 devices", "steps": "1. Spin on A\n2. Check B", "expected": "Independent results. No shared state", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet=0 rejected", "pre": "Idle", "steps": "1. Enter 0\n2. Click Spin", "expected": "No action. No spin", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify bet exceeds max", "pre": "Idle", "steps": "1. Enter 99999\n2. Click Spin", "expected": "Rejected or capped. Error", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Responsive", "title": "Verify wheel on small screen", "pre": "320px viewport", "steps": "1. Open game\n2. Observe wheel", "expected": "Wheel scales. Pointer visible. Spin button accessible. Segments readable", "priority": "Medium", "severity": "Medium"},
]

# ============================================================
# LOTTERY — Core Test Cases (ticket purchase, draw)
# ============================================================

LOTTERY_CORE = [
    {"module": "Game Engine", "feature": "Ticket Purchase", "title": "Verify buying a lottery ticket", "pre": "Balance=100, ticket=5", "steps": "1. Enter 5\n2. Click Buy Ticket", "expected": "5 deducted. Ticket number generated (6-digit). Countdown shown. Ticket saved", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Ticket Number", "title": "Verify unique 6-digit ticket number", "pre": "Ticket purchased", "steps": "1. Check ticket number", "expected": "6-digit number (100000-999999). Unique per round. Displayed on ticket card", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Prize Pool", "title": "Verify prize pool displays and updates", "pre": "Multiple participants", "steps": "1. Observe prize pool\n2. Buy ticket\n3. Check again", "expected": "Prize pool increases with each ticket. Updates live. Shows formatted amount", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Participant Count", "title": "Verify participant count", "pre": "Game loaded", "steps": "1. Observe participants field", "expected": "Shows total participants. Increases as players join. Includes simulated players", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Draw", "title": "Verify lottery draw execution", "pre": "Round timer expired", "steps": "1. Wait for draw time\n2. Observe draw", "expected": "Ticket flip animation. Winner revealed. Prize distributed. Results shown", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Win/Loss", "title": "Verify win shows prize, loss shows 'حظ أوفر'", "pre": "Draw completed", "steps": "1. Check result", "expected": "Win: '🎉 ربحت X!' + confetti. Loss: '💔 حظ أوفر'. Balance updates on win", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Ticket Flip", "title": "Verify ticket card flip animation", "pre": "Draw starting", "steps": "1. Observe ticket card during draw", "expected": "Card flips (rotateY 360°). Scratch overlay reveals. Number shown", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Auto-Buy", "title": "Verify auto-buy for N tickets", "pre": "Auto-buy=5, ticket=5", "steps": "1. Enable auto-buy\n2. Wait for 5 rounds", "expected": "5 tickets bought automatically. Counter decrements. Stops at 0", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal", "pre": "Balance=0", "steps": "1. Enter 5\n2. Click Buy", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Multiple Tickets", "title": "Verify buying multiple tickets in same round", "pre": "Balance=100", "steps": "1. Buy ticket for 5\n2. Buy another for 10", "expected": "Both tickets saved. Ticket count shows 2. 15 deducted total", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Ticket Card", "title": "Verify ticket card design", "pre": "Ticket purchased", "steps": "1. Observe ticket card", "expected": "Gold border. Ticket number large. Status 'لم تشترِ بعد' or 'تذاكرك: X'", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Countdown", "title": "Verify countdown before draw", "pre": "Tickets purchased", "steps": "1. Observe countdown", "expected": "3-2-1 countdown. Draw starts after. Sound tick", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "History Bar", "title": "Verify history after each draw", "pre": "Draw completed", "steps": "1. Observe history", "expected": "Result added. Green for win, red for loss", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Quick Stats", "title": "Verify stats panel", "pre": "Multiple draws", "steps": "1. Observe stats", "expected": "Best, draws, tickets count. Updates live", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "Provably Fair", "feature": "Seed Hash", "title": "Verify seed hash shown before draw", "pre": "Tickets sold, draw pending", "steps": "1. Check for seed commitment", "expected": "seed_hash displayed before draw. SHA256 commitment scheme", "priority": "Critical", "severity": "Critical"},
    {"module": "Provably Fair", "feature": "Server Seed Reveal", "title": "Verify server seed after draw", "pre": "Draw completed", "steps": "1. Check for revealed seed", "expected": "server_seed revealed. SHA256(server_seed) matches seed_hash", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify buying ticket at draw boundary", "pre": "Draw about to start", "steps": "1. Click Buy at exact draw time", "expected": "Server timestamp decides. If before: accepted. If after: rejected for next round", "priority": "Critical", "severity": "Critical"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance after win and loss", "pre": "Balance=100, ticket=5", "steps": "1. Win prize=50\n2. Lose", "expected": "After win: 145 (green). After loss: 95 (already deducted)", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "Server Error", "title": "Verify behavior on 500 during draw", "pre": "Draw in progress", "steps": "1. Simulate 500\n2. Observe", "expected": "Error toast. Result syncs on reconnect. No stuck state", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Independence", "title": "Verify per-player tickets", "pre": "2 devices", "steps": "1. Buy on A\n2. Check B", "expected": "Independent tickets. No shared ticket state", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify ticket=0 rejected", "pre": "Idle", "steps": "1. Enter 0\n2. Click Buy", "expected": "No action. No ticket", "priority": "Medium", "severity": "Medium"},
    {"module": "Betting", "feature": "Edge Case", "title": "Verify ticket below minimum rejected", "pre": "Idle, min=5", "steps": "1. Enter 1\n2. Click Buy", "expected": "Rejected. 'الحد الأدنى: 5' message", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify no participants in round", "pre": "Draw time reached, 0 tickets", "steps": "1. Wait for draw\n2. Observe", "expected": "Draw executes. No winner. Prize pool carries over or resets", "priority": "Low", "severity": "Low"},
    {"module": "UI", "feature": "Responsive", "title": "Verify on small screen", "pre": "320px viewport", "steps": "1. Open game\n2. Observe", "expected": "Ticket card fits. Buy button accessible. Prize pool readable", "priority": "Medium", "severity": "Medium"},
]

# ============================================================
# SNATCH (Gift) — Core Test Cases (catch-the-gift)
# ============================================================

SNATCH_CORE = [
    {"module": "Game Engine", "feature": "Gift Spawn", "title": "Verify gift boxes appear at random positions", "pre": "Game started", "steps": "1. Observe game area\n2. Wait for gifts to spawn", "expected": "Gift boxes appear at random positions. 4 types: golden, flash, bomb, decoy. 7 colors", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Catch Gift", "title": "Verify catching a golden gift", "pre": "Golden gift visible", "steps": "1. Tap golden gift", "expected": "Gift caught. Score increases. Confetti animation. Sound. Haptic feedback. Gift disappears", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Bomb", "title": "Verify tapping a bomb deducts a life", "pre": "3 lives, bomb visible", "steps": "1. Tap bomb gift", "expected": "Bomb explodes. -1 life. Penalty sound. Haptic heavy. Lives counter decrements", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Lives System", "title": "Verify game ends at 0 lives", "pre": "1 life remaining", "steps": "1. Tap a bomb", "expected": "Game over. Result screen shows all caught gifts. Score final. 'انتهت اللعبة'", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Flash Gift", "title": "Verify flash gift disappears quickly", "pre": "Flash gift spawns", "steps": "1. Observe flash gift\n2. Try to catch", "expected": "Flash gift stays 0.3-0.7s then disappears. 'فوتتك!' tease text. Very hard to catch", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Decoy Gift", "title": "Verify decoy moves position mid-appear", "pre": "Decoy spawns", "steps": "1. Observe decoy gift", "expected": "Gift appears, then moves to new position. 'خداع!' tease text. If caught: counts", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Score", "title": "Verify score tracking", "pre": "Game in progress", "steps": "1. Catch 3 gifts\n2. Check score", "expected": "Score = 3. Displayed prominently. Each catch adds +1", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Result Screen", "title": "Verify result screen after game ends", "pre": "Game over (0 lives)", "steps": "1. Observe result screen", "expected": "Shows all caught gifts. Total score. Telegram MainButton for claiming prizes", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Countdown", "title": "Verify 3-2-1 countdown before game start", "pre": "Game starting", "steps": "1. Observe countdown", "expected": "3-2-1 displayed. Game starts after. Sound tick. Haptic light", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Sound", "title": "Verify sounds for each event", "pre": "Sound ON", "steps": "1. Catch golden\n2. Tap bomb\n3. Miss gift", "expected": "Catch: chime. Bomb: explosion. Miss: tease sound. Toggle works", "priority": "Low", "severity": "Low"},
    {"module": "Game Engine", "feature": "Haptics", "title": "Verify haptic feedback", "pre": "TMA on mobile", "steps": "1. Catch gift\n2. Tap bomb", "expected": "Catch: light haptic. Bomb: heavy haptic. Via Telegram HapticFeedback API", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Tease Texts", "title": "Verify floating tease texts appear", "pre": "Gift missed or decoy", "steps": "1. Miss a gift\n2. Observe", "expected": "'فوتتك!' or 'خداع!' or 'اختطفتها!' floats up and fades", "priority": "Low", "severity": "Low"},
    {"module": "Betting", "feature": "Place Bet", "title": "Verify bet starts game", "pre": "Balance=100, bet=10", "steps": "1. Enter 10\n2. Click Start", "expected": "10 deducted. Countdown starts. Game begins after countdown", "priority": "Critical", "severity": "Critical"},
    {"module": "Betting", "feature": "Insufficient Balance", "title": "Verify deposit modal", "pre": "Balance=0", "steps": "1. Enter 10\n2. Click Start", "expected": "VEX deposit modal appears", "priority": "High", "severity": "High"},
    {"module": "Game Engine", "feature": "Prize Claim", "title": "Verify claiming prizes via MainButton", "pre": "Game ended", "steps": "1. Click MainButton 'Claim Prizes'", "expected": "sendData sends caught gifts to bot. Bot processes prizes. Frozen balance added", "priority": "High", "severity": "High"},
    {"module": "UI", "feature": "Connection Indicator", "title": "Verify connection status", "pre": "Game loaded", "steps": "1. Disconnect\n2. Reconnect", "expected": "🟢→🔴→🟡→🟢", "priority": "High", "severity": "High"},
    {"module": "Betting", "feature": "Race Condition", "title": "Verify double-clicking Start", "pre": "Idle", "steps": "1. Click Start 3x in 100ms", "expected": "Only 1 game starts. 1 deduction. Button disables after 1st", "priority": "Critical", "severity": "Critical"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify catching gift at exact despawn time", "pre": "Gift about to despawn", "steps": "1. Tap at exact despawn frame", "expected": "Server determines result. If caught before despawn: counts. If after: miss", "priority": "High", "severity": "High"},
    {"module": "Wallet", "feature": "Balance Sync", "title": "Verify balance after game", "pre": "Balance=100, bet=10", "steps": "1. Play game\n2. Catch 5 gifts\n3. Claim prizes", "expected": "After bet: 90. After claim: 90 + prize value. Balance matches server", "priority": "High", "severity": "High"},
    {"module": "Error Handling", "feature": "Server Error", "title": "Verify behavior on 500 during game", "pre": "Game in progress", "steps": "1. Simulate 500\n2. Observe", "expected": "Game continues (client-side). Prizes sync on reconnect. No data loss", "priority": "High", "severity": "High"},
    {"module": "Multi-Device", "feature": "Independence", "title": "Verify per-player game", "pre": "2 devices", "steps": "1. Start on A\n2. Check B", "expected": "Independent games. No shared gifts", "priority": "Medium", "severity": "Medium"},
    {"module": "UI", "feature": "Responsive", "title": "Verify on small screen", "pre": "320px viewport", "steps": "1. Open game\n2. Play", "expected": "Gifts spawn within visible area. Catchable on small screen. No overflow", "priority": "Medium", "severity": "Medium"},
    {"module": "Game Engine", "feature": "Edge Case", "title": "Verify bet=0 rejected", "pre": "Idle", "steps": "1. Enter 0\n2. Click Start", "expected": "No action. No game", "priority": "Medium", "severity": "Medium"},
]

def generate_excel(game_name, game_prefix, core_cases, matrices, output_dir="qa_exports"):
    """Generate Excel file with Core + Matrix-expanded test cases."""
    os.makedirs(output_dir, exist_ok=True)
    final_cases = []
    test_counter = 1

    # 1. Add Core Functional cases
    for case in core_cases:
        final_cases.append({
            "Test ID": f"{game_prefix}-FUNC-{test_counter:03d}",
            "Module": case["module"],
            "Feature": case["feature"],
            "Test Title": case["title"],
            "Preconditions": case["pre"],
            "Test Steps": case["steps"],
            "Expected Result": case["expected"],
            "Priority": case["priority"],
            "Severity": case["severity"],
            "Test Type": case.get("type_override", "Functional"),
            "Status": "Draft",
            "Notes": "",
        })
        test_counter += 1

    # 2. Expand each Core case × each Matrix
    for matrix in matrices:
        for case in core_cases:
            # Skip illogical combos
            if case["module"] == "Provably Fair" and matrix["type"] == "Performance":
                continue
            if case["module"] == "Database" and matrix["type"] in ("UI", "Telegram Mini App"):
                continue

            cat_prefix = matrix["type"][:4].upper()
            final_cases.append({
                "Test ID": f"{game_prefix}-{cat_prefix}-{test_counter:03d}",
                "Module": case["module"],
                "Feature": case["feature"],
                "Test Title": f"{case['title']} {matrix['suffix']}",
                "Preconditions": case["pre"],
                "Test Steps": f"{case['steps']}\n--- Matrix ---\n{matrix['steps']}",
                "Expected Result": f"{matrix['expected']}\n(Original: {case['expected']})",
                "Priority": matrix["priority"],
                "Severity": matrix["severity"],
                "Test Type": matrix["type"],
                "Status": "Draft",
                "Notes": "Auto-generated via Matrix Approach",
            })
            test_counter += 1

    # 3. Write to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{game_name} QA"

    # Headers
    headers = ["Test ID", "Module", "Feature", "Test Title", "Preconditions",
               "Test Steps", "Expected Result", "Priority", "Severity",
               "Test Type", "Status", "Notes"]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    priority_colors = {
        "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "High": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Medium": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Low": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    for row_idx, case in enumerate(final_cases, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if header == "Priority":
                fill = priority_colors.get(case.get("Priority", ""), None)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)

    # Column widths
    widths = {"A": 16, "B": 14, "C": 16, "D": 40, "E": 28, "F": 36, "G": 40, "H": 10, "I": 10, "J": 16, "K": 8, "L": 20}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(final_cases) + 1}"

    filepath = os.path.join(output_dir, f"{game_name}_QA_TestCases.xlsx")
    wb.save(filepath)
    print(f"[OK] {game_name}: {len(final_cases)} test cases -> {filepath}")
    return len(final_cases)


# ============================================================
# RUN — Generate for all games
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("QA Matrix Generator — Boterx VEX Games")
    print("=" * 60)

    total = 0

    # Aviator
    total += generate_excel("Aviator", "AVI", AVIATOR_CORE, ALL_MATRICES)

    # Crash
    total += generate_excel("Crash", "CRS", CRASH_CORE, ALL_MATRICES)

    # Mines
    total += generate_excel("Mines", "MIN", MINES_CORE, ALL_MATRICES)

    # Plinko
    total += generate_excel("Plinko", "PLK", PLINKO_CORE, ALL_MATRICES)

    # Lucky Wheel
    total += generate_excel("LuckyWheel", "WHE", WHEEL_CORE, ALL_MATRICES)

    # Lottery
    total += generate_excel("Lottery", "LOT", LOTTERY_CORE, ALL_MATRICES)

    # Snatch (Gift)
    total += generate_excel("Snatch", "SNC", SNATCH_CORE, ALL_MATRICES)

    print(f"\n{'=' * 60}")
    print(f"TOTAL: {total} test cases generated")
    print(f"Files in: qa_exports/")
    print(f"\nTo add more games: define CORE cases list + call generate_excel()")
    print(f"Pattern: {len(AVIATOR_CORE)} Core × {len(ALL_MATRICES)} Matrices = {len(AVIATOR_CORE) * len(ALL_MATRICES) + len(AVIATOR_CORE)} cases")
